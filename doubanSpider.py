#-*- coding: UTF-8 -*-

import sys
import time
import urllib
import urllib.parse
import requests
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook
from imp import reload
import re
import configparser
import pymysql

connection = pymysql.connect(host='127.0.0.1',
                             port=3306,
                             user='douban',
                             password='123456',
                             db='douban',
                             charset='utf8')


cp = configparser.SafeConfigParser()
cp.read('.passwd')
un=cp.get('douban', 'username')
pw=cp.get('douban', 'password')

reload(sys)
#sys.setdefaultencoding('utf8')

s = requests.Session()


#Some User Agents
hds=[{'User-Agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36'}]

def login_douban():
    """
    登录豆瓣
    :return:
    """
    # 登录URL
    login_url = 'https://accounts.douban.com/login'
    # 请求头
    data = {
        'source':'index_nav',
        'redir':'https://www.douban.com/',
        'form_email':"un",
        'form_password':"pw",
        'login':'登录'
    }
    try:
        r = s.post(login_url, headers=hds[0], data=data)
        r.raise_for_status()
    except:
        print('登录请求失败')
        return 0
    # 打印请求结果
    return 1


def book_spider(book_tag):
    page_num=0;
    try_times=0
    book_list=[]

    while(1):
        #url='http://www.douban.com/tag/%E5%B0%8F%E8%AF%B4/book?start=0' # For Test
        url='http://www.douban.com/tag/'+urllib.parse.quote(book_tag)+'/book?start='+str(page_num*15)
        time.sleep(np.random.rand()*5)

        #Last Version
        try:
#            req = urllib.request.Request(url, headers=hds[page_num%len(hds)])
#            source_code = urllib.request.urlopen(req).read()
#            plain_text=str(source_code)
            r = s.get(url, headers=hds[page_num%len(hds)])
            plain_text=r.text
        except (requests.exceptions.HTTPError, requests.exceptions.Timeout) as e:
            print (e)
            continue

        ##Previous Version, IP is easy to be Forbidden
        #source_code = requests.get(url)
        #plain_text = source_code.text

        soup = BeautifulSoup(plain_text, "html.parser")
        list_soup = soup.find('div', {'class': 'mod book-list'})

        try_times+=1;
        if list_soup==None and try_times<200:
            continue
        elif list_soup==None or len(list_soup)<=1:
            break # Break when no informatoin got after 200 times requesting

        for book_info in list_soup.findAll('dd'):
            title = book_info.find('a', {'class':'title'}).string.strip()
            douban_id = int(book_info.a.get('href').split('/')[4])
            desc = book_info.find('div', {'class':'desc'}).string.strip()
            desc_list = desc.split('/')
            book_url = book_info.find('a', {'class':'title'}).get('href')

            try:
                author_info = str(desc_list[0:-3])
            except:
                author_info ='暂无'
            try:
                book_price = float(re.findall(r'(\d+\.\d+|\d+)', (desc_list[-1]))[0])
            except:
                book_price = 0.0
            try:
                press_date = str(desc_list[-2])
            except:
                press_date = '0001-01-01'
            try:
                pub_press = str(desc_list[-3])
            except:
                pub_press = '暂无'
            try:
                rating = float(book_info.find('span', {'class':'rating_nums'}).string.strip())
            except:
                rating=float(0.0)
            try:
                #people_num = book_info.findAll('span')[2].string.strip()
                people_snum = get_people_num(book_url)
                people_num = int(people_snum.strip('人评价'))
            except:
                people_num=int(0)

            #book_title,book_douban_id,book_rate,book_author,book_rate_user,book_press,book_press_date,book_price
            book_single=(title,douban_id,rating,author_info,people_num,pub_press,press_date,book_price,book_tag)
            book_list.append(book_single)
            try_times=0 #set 0 when got valid information
        page_num+=1
        print ('Downloading Information From Page %d' % page_num)
#        print (type(book_list))
#    tuple_book_list=tuple(book_list)
    return book_list



def get_people_num(url):
    #url='http://book.douban.com/subject/6082808/?from=tag_all' # For Test
    try:
        req = urllib.request.Request(url, headers=hds[np.random.randint(0,len(hds))])
        source_code = urllib.request.urlopen(req).read()
        plain_text=str(source_code)
    except (urllib.request.HTTPError, urllib.request.URLError) as e:
        print (e)
    soup = BeautifulSoup(plain_text, "html.parser")
    people_num=soup.find('div',{'class':'rating_sum'}).findAll('span')[1].string.strip()
    return people_num


def do_spider(book_tag_lists):
#    book_lists=[]
    cursor = connection.cursor()
    for book_tag in book_tag_lists:
        book_list=book_spider(book_tag)
#        book_lists=book_lists.append(book_list)
#        tuple_book_lists=tuple(book_lists)
        print(len(book_list[0]))
        print(book_list[0])
#        book_list=sorted(book_list,key=lambda x:x[1],reverse=True)
        cursor = connection.cursor()
        cursor.executemany(
            'INSERT INTO bookinfo (book_title, book_douban_id,book_rate,book_author,book_rate_user,book_press,book_press_date,book_price,book_tag) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)', book_list)
#        SQL='INSERT INTO bookinfo ("book_title", "book_douban_id","book_rate","book_author","book_rate_user","book_press","book_press_date","book_price","book_tag") VALUES ("%s", %d, %f, "%s", %d, "%s", "%s", %f, "%s")' % book_list[0]
#        print(SQL)
        connection.commit()
#        print(type(book_list))
#        book_lists.append(book_list)
#    return book_lists

#def do_spider(book_tag_lists):
#    book_lists=[]
#    for book_tag in book_tag_lists:
#        book_list=book_spider(book_tag)
#        book_list=sorted(book_list,key=lambda x:x[1],reverse=True)
#        book_lists.append(book_list)
#    return book_lists
#
#
#def print_book_lists_excel(book_lists,book_tag_lists):
#    wb=Workbook(write_only = True)
#    ws=[]
#    for i in range(len(book_tag_lists)):
#        print(type(book_tag_lists[i]))
#        ws.append(wb.create_sheet(title=book_tag_lists[i].encode('utf-8').decode('unicode_escape'))) #utf8->unicode
#    for i in range(len(book_tag_lists)):
#        ws[i].append(['序号','书名','评分','评价人数','作者','出版社'])
#        count=1
#        for bl in book_lists[i]:
#            ws[i].append([count,bl[0],float(bl[1]),int(bl[2]),bl[3],bl[4]])
#            count+=1
#    save_path='book_list'
#    for i in range(len(book_tag_lists)):
#        save_path+=('-'+book_tag_lists[i].encode('utf-8').decode('unicode_escape'))
#    save_path+='.xlsx'
#    wb.save(save_path)




if __name__=='__main__':
    #book_tag_lists = ['心理','判断与决策','算法','数据结构','经济','历史']
    #book_tag_lists = ['传记','哲学','编程','创业','理财','社会学','佛教']
    #book_tag_lists = ['思想','科技','科学','web','股票','爱情','两性']
    #book_tag_lists = ['计算机','机器学习','linux','android','数据库','互联网']
    #book_tag_lists = ['数学']
    #book_tag_lists = ['摄影','设计','音乐','旅行','教育','成长','情感','育儿','健康','养生']
    #book_tag_lists = ['商业','理财','管理']
    #book_tag_lists = ['名著']
    #book_tag_lists = ['科普','经典','生活','心灵','文学']
    #book_tag_lists = ['科幻','思维','金融']
    #book_tag_lists = ['个人管理','时间管理','投资','文化','宗教']
    if login_douban():
        book_tag_lists = ['linux']
        do_spider(book_tag_lists)
#        print_book_lists_excel(book_lists,book_tag_lists)
